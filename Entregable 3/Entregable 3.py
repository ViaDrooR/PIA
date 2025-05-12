from openpyxl import Workbook
from statistics import mean, median, stdev

def promedio (lista, clave):
    prom = mean(lista)
    print(f"El promedio de {clave} es: {prom:.2f}")

def mediana (lista, clave):
    mediana = median(lista)
    print(f"La mediana de {clave} es: {mediana:.2f}")

def desviacion_est (lista, clave):
    desv = stdev(lista)
    print(f"La desviación estándar de {clave} es: {desv:.2f}")

lista_asteroides = []
valores = []
with open("asteroides_2024.txt", "r", encoding="utf-8") as archivo:
    for linea in archivo:
        if linea != "\n":
            if ":" in linea:
                indice = linea.index(":")
                value = linea[indice+2:-1]
                valores.append(value)
        else:
            asteroide = {
                "Fecha": valores[0],
                "Nombre": valores[1],
                "Es peligroso": valores[2],
                "Tamaño (m)": float(valores[3]),
                "Velocidad (km/h)": float(valores[4]),
                "Distancia a la Tierra (km)": float(valores[5]),
            }
            lista_asteroides.append(asteroide)
            valores.clear()

if lista_asteroides:
    libro=Workbook()
    hoja=libro.active
    hoja.title="Asteroides"
    #Escribir encabezados
    encabezados=list(lista_asteroides[0].keys())
    columna=1 #Inicia el contador en 1
    for encabezado in encabezados:
        hoja.cell(row=1, column=columna, value=encabezado)
        columna+=1 #Aumenta el contador de columnas 

    #Escribir filas con datos
    fila_actual=2
    distancia = []
    tamanio = []
    velocidad = []
    for asteroide in lista_asteroides:
        col_actual=1  #Reinicia el contador de columnas en cada fila
        for clave in asteroide:
            hoja.cell(row=fila_actual, column=col_actual, value=asteroide[clave])
            if clave == "Distancia a la Tierra (km)":
                distancia.append(asteroide[clave])
                clave_dis = clave
            elif clave == "Tamaño (m)":
                tamanio.append(asteroide[clave])
                clave_tam = clave
            elif clave == "Velocidad (km/h)":
                velocidad.append(asteroide[clave])
                clave_vel = clave
            else:
                pass 
            col_actual+=1  #Aumenta el contador de columnas
        fila_actual += 1  #Aumenta la fila 
    libro.save("asteroides.xlsx")

promedio(distancia, clave_dis)
desviacion_est(distancia, clave_dis)
mediana(distancia, clave_dis)
print("---------------------------")
promedio(tamanio, clave_tam)
desviacion_est(tamanio, clave_tam)
mediana(tamanio, clave_tam)
print("---------------------------")
promedio(velocidad, clave_vel)
desviacion_est(velocidad, clave_vel)
mediana(velocidad, clave_vel)


