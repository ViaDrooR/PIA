from openpyxl import Workbook
from statistics import mean, median, stdev
import matplotlib.pyplot as plt


    #---Agregué esto para poder imprimir el promedio---
def promedio_show (lista, clave=""):
    prom_s=mean(lista)
    print(f"El promedio de {clave} es: {prom_s:.2f}")

def promedio (lista, clave=""):
    prom = mean(lista)
    return prom
    #print(f"El promedio de {clave} es: {prom:.2f}")

def mediana (lista, clave):
    mediana = median(lista)
    print(f"La mediana de {clave} es: {mediana:.2f}")

def desviacion_est (lista, clave):
    desv = stdev(lista)
    print(f"La desviación estándar de {clave} es: {desv:.2f}")

lista_asteroides = []
valores = []
with open("asteroides_2024.txt", "r", encoding="utf-8") as archivo:
    for linea in archivo:
        if linea != "\n":
            if ":" in linea:
                indice = linea.index(":")
                value = linea[indice+2:-1]
                valores.append(value)
        else:
            asteroide = {
                "Fecha": valores[0],
                "Nombre": valores[1],
                "Es peligroso": valores[2],
                "Tamaño (m)": float(valores[3]),
                "Velocidad (km/h)": float(valores[4]),
                "Distancia a la Tierra (km)": float(valores[5]),
            }
            lista_asteroides.append(asteroide)
            valores.clear()

if lista_asteroides:
    libro=Workbook()
    hoja=libro.active
    hoja.title="Asteroides"
    #Escribir encabezados
    encabezados=list(lista_asteroides[0].keys())
    columna=1 #Inicia el contador en 1
    for encabezado in encabezados:
        hoja.cell(row=1, column=columna, value=encabezado)
        columna+=1 #Aumenta el contador de columnas 

    #Escribir filas con datos
    fila_actual=2
    distancia = []
    tamanio = []
    velocidad = []
    for asteroide in lista_asteroides:
        col_actual=1  #Reinicia el contador de columnas en cada fila
        for clave in asteroide:
            hoja.cell(row=fila_actual, column=col_actual, value=asteroide[clave])
            if clave == "Distancia a la Tierra (km)":
                distancia.append(asteroide[clave])
                clave_dis = clave
            elif clave == "Tamaño (m)":
                tamanio.append(asteroide[clave])
                clave_tam = clave
            elif clave == "Velocidad (km/h)":
                velocidad.append(asteroide[clave])
                clave_vel = clave
            else:
                pass 
            col_actual+=1  #Aumenta el contador de columnas
        fila_actual += 1  #Aumenta la fila 
    libro.save("asteroides.xlsx")

promedio_show(distancia, clave_dis)
desviacion_est(distancia, clave_dis)
mediana(distancia, clave_dis)
print("---------------------------")
promedio_show(tamanio, clave_tam)
desviacion_est(tamanio, clave_tam)
mediana(tamanio, clave_tam)
print("---------------------------")
promedio_show(velocidad, clave_vel)
desviacion_est(velocidad, clave_vel)
mediana(velocidad, clave_vel)

#NOTA: voy a usar plt.figure() la cual hace que se muestren todas las graficas al mismo tiempo


#Creación de gráficas
    #---Edité esto para que sea más automático la extracción de datos---
velocidad_por_mes={}
for ast in lista_asteroides:
    fecha=ast["Fecha"]           #ejemplo: "2024-03-21"
    mes=fecha[5:7]               #extrae "03"
    if mes not in velocidad_por_mes:
        velocidad_por_mes[mes]=[]  #Si no existe el mes en el diccionario, lo crea
    velocidad_por_mes[mes].append(ast["Velocidad (km/h)"])   #Agrega el dato del asteroide al mes correspondiente
  #Creación de la lista ordenada por mes
meses=["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
meses_numeros=["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
y_label=[]
for num in meses_numeros:
    vel_mes=velocidad_por_mes[num]#Obtener los datos del mes
    y_label.append(promedio(vel_mes)) #Agrega el promedio del mes a la lista
    
#Creación de gráfica de barras
    
#Primera gráfica
plt.figure()
plt.bar(meses, y_label)
plt.axis(ymin=40000, ymax=48000)
plt.title("Comparación de promedios de velocidades por mes, del año 2024")
plt.ylabel("Velocidad (m/s)")
plt.xlabel("Meses")

#Creación de gráfica plot
    #---También este---
tamanio_por_mes={}
for ast in lista_asteroides:
    fecha=ast["Fecha"]           
    mes=fecha[5:7]
    if mes not in tamanio_por_mes:
        tamanio_por_mes[mes]=[]
    tamanio_por_mes[mes].append(ast["Tamaño (m)"])
  #Creación de la lista ordenada por mes
meses_numeros=["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]
y_label_tamanio=[]  
for num in meses_numeros:
    tam_mes=tamanio_por_mes[num]  
    y_label_tamanio.append(promedio(tam_mes))  #Agrega el promedio del mes a la lista
    #---Y este---
mas_grande=[]
mas_pequenio=[]
for num in meses_numeros:
    tam_mes=tamanio_por_mes[num]
    tam_mes.sort()
    mas_grande.append(tam_mes[-1])
    mas_pequenio.append(tam_mes[0])
    
#Segunda gráfica
plt.figure()
plt.plot(meses, mas_pequenio, marker="o")
plt.title("Comparación de los asteroides más pequeños por mes, del año 2024")
plt.ylabel("Tamaño (m)")
plt.xlabel("Meses")
#Tercera gráfica
plt.figure()
plt.plot(meses, mas_grande, marker="o")
plt.title("Comparación de los asteroides más grandes por mes, del año 2024")
plt.ylabel("Tamaño (m)")
plt.xlabel("Meses")


#Diagramas de dispersión, Tamaño / Valocidad
plt.figure()
plt.scatter(tamanio, velocidad)
plt.title("Diagrama de dispersión: Tamaño vs Velocidad de Asteroides")
plt.xlabel("Tamaño (m)")
plt.ylabel("Velocidad (km/h)")


#Gráficos de pastel. Asteroides por mes
#Cantidad de asteroides por mes
conteo_por_mes={}
for ast in lista_asteroides:
    mes=ast["Fecha"][5:7]
    if mes not in conteo_por_mes:
        conteo_por_mes[mes]=0
    conteo_por_mes[mes]+=1
cantidad_asteroides=[]
for num in meses_numeros:
    cantidad_asteroides.append(conteo_por_mes.get(num, 0))
#Gráfico de pastel
plt.figure()
plt.pie(cantidad_asteroides, labels=meses, autopct='%1.1f%%') #NOTA: autopct='%1.1f%%' Esto muestra el porcentaje con un decimal
plt.title("Distribución de asteroides por mes (2024)")


#Gráficos de pastel. Peligrosidad
conteo_peligrosidad={}
conteo_peligrosidad["Peligroso"]=0
conteo_peligrosidad["No peligroso"]=0

for ast in lista_asteroides:
    if ast["Es peligroso"]=="True":
        conteo_peligrosidad["Peligroso"]+=1
    else:
        conteo_peligrosidad["No peligroso"]+=1
valores=list(conteo_peligrosidad.values())
etiquetas=list(conteo_peligrosidad.keys())
plt.figure()
plt.pie(valores, labels=etiquetas, autopct='%1.1f%%')
plt.title("Distribución de asteroides por peligrosidad (2024)")

plt.show()


print("FIN")
